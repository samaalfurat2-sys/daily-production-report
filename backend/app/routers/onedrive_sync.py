"""
OneDrive sync router.

Endpoints
---------
GET  /onedrive/status          – check if OneDrive is configured
GET  /onedrive/setup           – start device-code auth flow
POST /onedrive/setup/complete  – complete auth and save refresh token
POST /onedrive/backup/db       – upload SQLite DB file to OneDrive
POST /onedrive/export/shifts   – export shift reports as Excel to OneDrive
POST /onedrive/export/inventory – export inventory/stock as Excel to OneDrive
GET  /onedrive/files           – list files in OneDrive folder
POST /onedrive/sync/all        – run all exports + DB backup in one call
"""

import io
import os
import logging
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app import models
from app.deps import get_db, require_role, get_current_user
from app import onedrive_client as od

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/onedrive", tags=["OneDrive"])

_FOLDER = os.environ.get("ONEDRIVE_FOLDER", "ProductionReports")

# ── device-code state (in-memory, single-instance) ────────────────────────────
_pending_device_code: dict = {}


# ── Schemas ───────────────────────────────────────────────────────────────────

class SetupStartResponse(BaseModel):
    user_code: str
    verification_uri: str
    message: str
    device_code: str
    expires_in: int


class SetupCompleteRequest(BaseModel):
    device_code: str


class SetupCompleteResponse(BaseModel):
    ok: bool
    message: str
    refresh_token_preview: str   # first 8 chars only, for confirmation


class StatusResponse(BaseModel):
    configured: bool
    folder: str
    message: str


class SyncResult(BaseModel):
    ok: bool
    db_backup: Optional[str] = None
    shifts_excel: Optional[str] = None
    inventory_excel: Optional[str] = None
    errors: List[str] = []


# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_onedrive():
    if not od.is_configured():
        raise HTTPException(
            status_code=503,
            detail="OneDrive is not configured. Call GET /onedrive/setup first.",
        )


def _build_shifts_excel(db: Session) -> bytes:
    """Export all approved/locked shifts with sub-reports to Excel bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Add it to requirements.txt.",
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Shift Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Shifts"
    header = ["ID", "Date", "Shift Code", "Status", "Created At",
              "Submitted At", "Approved At", "Notes"]
    _excel_header_row(ws, header)
    shifts = db.query(models.ShiftRecord).order_by(
        models.ShiftRecord.report_date.desc()
    ).all()
    for s in shifts:
        ws.append([
            str(s.id),
            str(s.report_date),
            s.shift_code,
            s.status,
            str(s.created_at)[:19] if s.created_at else "",
            str(s.submitted_at)[:19] if s.submitted_at else "",
            str(s.approved_at)[:19] if s.approved_at else "",
            s.notes or "",
        ])
    _autofit(ws)

    # ── Sheet 2: Blow Reports ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Blow")
    blow_header = [
        "Shift ID", "Date", "Shift Code",
        "Preforms/Carton", "Prev Cartons", "Received Cartons",
        "Next Cartons", "Product Cartons",
        "Waste Preforms (pcs)", "Waste Scrap (pcs)", "Waste Bottles (pcs)",
        "Counter Value",
    ]
    _excel_header_row(ws2, blow_header)
    for s in shifts:
        if s.blow:
            b = s.blow
            ws2.append([
                str(s.id), str(s.report_date), s.shift_code,
                b.preforms_per_carton,
                float(b.prev_cartons) if b.prev_cartons is not None else "",
                float(b.received_cartons) if b.received_cartons is not None else "",
                float(b.next_cartons) if b.next_cartons is not None else "",
                float(b.product_cartons) if b.product_cartons is not None else "",
                b.waste_preforms_pcs or "",
                b.waste_scrap_pcs or "",
                b.waste_bottles_pcs or "",
                b.counter_value or "",
            ])
    _autofit(ws2)

    # ── Sheet 3: Filling Reports ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Filling")
    fill_header = [
        "Shift ID", "Date", "Shift Code",
        "Caps/Carton", "Prev Cartons", "Received Cartons", "Next Cartons",
        "Waste Caps (pcs)", "Waste Scrap (pcs)", "Waste Bottles (pcs)",
        "Counter Value",
    ]
    _excel_header_row(ws3, fill_header)
    for s in shifts:
        if s.filling:
            f = s.filling
            ws3.append([
                str(s.id), str(s.report_date), s.shift_code,
                f.caps_per_carton,
                float(f.prev_cartons) if f.prev_cartons is not None else "",
                float(f.received_cartons) if f.received_cartons is not None else "",
                float(f.next_cartons) if f.next_cartons is not None else "",
                f.waste_caps_pcs or "",
                f.waste_scrap_pcs or "",
                f.waste_bottles_pcs or "",
                f.counter_value or "",
            ])
    _autofit(ws3)

    # ── Sheet 4: Label Reports ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Label")
    lbl_header = [
        "Shift ID", "Date", "Shift Code",
        "Labels/Roll", "Prev Rolls", "Received Rolls", "Next Rolls",
        "Waste (grams)",
    ]
    _excel_header_row(ws4, lbl_header)
    for s in shifts:
        if s.label:
            l = s.label
            ws4.append([
                str(s.id), str(s.report_date), s.shift_code,
                l.labels_per_roll,
                float(l.prev_rolls) if l.prev_rolls is not None else "",
                float(l.received_rolls) if l.received_rolls is not None else "",
                float(l.next_rolls) if l.next_rolls is not None else "",
                float(l.waste_grams) if l.waste_grams is not None else "",
            ])
    _autofit(ws4)

    # ── Sheet 5: Shrink Reports ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Shrink")
    shrink_header = [
        "Shift ID", "Date", "Shift Code",
        "kg/Roll", "Prev Rolls", "Received Rolls", "Next Rolls",
        "Waste (kg)", "Screen Counter",
    ]
    _excel_header_row(ws5, shrink_header)
    for s in shifts:
        if s.shrink:
            sh = s.shrink
            ws5.append([
                str(s.id), str(s.report_date), s.shift_code,
                float(sh.kg_per_roll) if sh.kg_per_roll is not None else "",
                float(sh.prev_rolls) if sh.prev_rolls is not None else "",
                float(sh.received_rolls) if sh.received_rolls is not None else "",
                float(sh.next_rolls) if sh.next_rolls is not None else "",
                float(sh.waste_kg) if sh.waste_kg is not None else "",
                sh.screen_counter or "",
            ])
    _autofit(ws5)

    # ── Sheet 6: Diesel Reports ───────────────────────────────────────────────
    ws6 = wb.create_sheet("Diesel")
    die_header = [
        "Shift ID", "Date", "Shift Code",
        "Gen1 Total Reading", "Gen1 Consumed",
        "Gen2 Total Reading", "Gen2 Consumed",
        "Main Tank Received",
    ]
    _excel_header_row(ws6, die_header)
    for s in shifts:
        if s.diesel:
            d = s.diesel
            ws6.append([
                str(s.id), str(s.report_date), s.shift_code,
                float(d.generator1_total_reading) if d.generator1_total_reading is not None else "",
                float(d.generator1_consumed) if d.generator1_consumed is not None else "",
                float(d.generator2_total_reading) if d.generator2_total_reading is not None else "",
                float(d.generator2_consumed) if d.generator2_consumed is not None else "",
                float(d.main_tank_received) if d.main_tank_received is not None else "",
            ])
    _autofit(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_inventory_excel(db: Session) -> bytes:
    """Export warehouses, items and recent transactions to Excel bytes."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Add it to requirements.txt.",
        )

    wb = openpyxl.Workbook()

    # ── Stock on hand (computed from transactions) ────────────────────────────
    ws = wb.active
    ws.title = "Stock"
    _excel_header_row(ws, ["Warehouse Code", "Warehouse Name", "Item Code",
                            "Item Name", "UOM", "Balance Qty"])

    # Single query: fetch all transactions joined with warehouse + item info,
    # then compute balances in memory to avoid N×M per-row queries.
    all_txns = (
        db.query(models.InventoryTransaction)
        .join(models.Warehouse,
              models.InventoryTransaction.warehouse_id == models.Warehouse.id)
        .join(models.InventoryItem,
              models.InventoryTransaction.item_id == models.InventoryItem.id)
        .filter(
            models.Warehouse.is_active,
            models.InventoryItem.is_active,
        )
        .all()
    )

    # Accumulate balance per (warehouse_id, item_id) pair.
    # RECEIVE and ADJUST are positive; ISSUE is negative.
    _POSITIVE_TYPES = {"RECEIVE", "ADJUST"}
    balances: dict = {}
    wh_meta: dict = {}
    item_meta: dict = {}
    for t in all_txns:
        key = (t.warehouse_id, t.item_id)
        sign = 1.0 if t.txn_type in _POSITIVE_TYPES else -1.0
        balances[key] = balances.get(key, 0.0) + sign * float(t.qty)
        wh_meta[t.warehouse_id] = t.warehouse
        item_meta[t.item_id] = t.item

    for (wh_id, item_id), balance in sorted(
        balances.items(), key=lambda kv: (str(kv[0][0]), str(kv[0][1]))
    ):
        wh = wh_meta[wh_id]
        item = item_meta[item_id]
        ws.append([wh.code, wh.name_en, item.code, item.name_en,
                   item.uom, round(balance, 3)])
    _autofit(ws)

    # ── Transactions ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Transactions")
    _excel_header_row(ws2, ["Date", "Warehouse Code", "Warehouse Name",
                             "Item Code", "Item Name", "Type",
                             "Qty", "Reference Type", "Reference ID",
                             "Note", "Created At"])
    txns = (
        db.query(models.InventoryTransaction)
        .order_by(models.InventoryTransaction.created_at.desc())
        .limit(2000)
        .all()
    )
    for t in txns:
        wh = t.warehouse
        item = t.item
        ws2.append([
            str(t.txn_date),
            wh.code if wh else "",
            wh.name_en if wh else "",
            item.code if item else "",
            item.name_en if item else "",
            t.txn_type,
            float(t.qty),
            t.reference_type or "",
            t.reference_id or "",
            t.note or "",
            str(t.created_at)[:19] if t.created_at else "",
        ])
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_header_row(ws, headers: list):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/status", response_model=StatusResponse)
def onedrive_status():
    """Check whether OneDrive is configured and ready."""
    configured = od.is_configured()
    return StatusResponse(
        configured=configured,
        folder=_FOLDER,
        message="Ready — OneDrive sync is active." if configured
                else "Not configured. Visit GET /onedrive/setup to authenticate.",
    )


@router.get("/setup", response_model=SetupStartResponse)
def onedrive_setup_start():
    """
    Start Microsoft device-code authentication.
    Returns a user_code the user must enter at https://microsoft.com/devicelogin.
    Then call POST /onedrive/setup/complete with the device_code.
    """
    global _pending_device_code
    data = od.get_device_code()
    _pending_device_code = data
    return SetupStartResponse(**data)


@router.post("/setup/complete", response_model=SetupCompleteResponse)
def onedrive_setup_complete(body: SetupCompleteRequest):
    """
    Poll Microsoft until the user completes sign-in, then save the refresh token.
    Set the returned refresh_token as ONEDRIVE_REFRESH_TOKEN environment variable.
    """
    try:
        tokens = od.exchange_device_code(body.device_code)
        refresh = tokens["refresh_token"]
        os.environ["ONEDRIVE_REFRESH_TOKEN"] = refresh
        preview = refresh[:8] + "..." if len(refresh) > 8 else refresh
        return SetupCompleteResponse(
            ok=True,
            message=(
                "✅ OneDrive authenticated successfully!\n"
                f"Save this refresh token as ONEDRIVE_REFRESH_TOKEN env var:\n{refresh}"
            ),
            refresh_token_preview=preview,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/backup/db")
def backup_database(_: None = Depends(require_role("admin"))):
    """
    Upload the SQLite database file to OneDrive/<ONEDRIVE_FOLDER>/backups/.
    Only works when DATABASE_URL points to a local SQLite file.
    """
    _require_onedrive()
    db_url = os.environ.get("DATABASE_URL", "sqlite:///./production_app.db")
    if "sqlite" not in db_url:
        raise HTTPException(
            status_code=400,
            detail="DB backup to OneDrive only supported for SQLite databases.",
        )
    # Extract file path from URL  sqlite:///./production_app.db
    db_path = db_url.replace("sqlite:///", "").replace("sqlite://", "")
    if not os.path.exists(db_path):
        # Try common locations
        for candidate in ["./production_app.db", "production_app.db", "/app/production_app.db"]:
            if os.path.exists(candidate):
                db_path = candidate
                break
        else:
            raise HTTPException(status_code=404,
                                detail=f"SQLite file not found at {db_path}")

    with open(db_path, "rb") as f:
        content = f.read()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"production_app_{ts}.db"
    folder = f"{_FOLDER}/backups"
    web_url = od.upload_file(folder, filename, content, "application/octet-stream")
    return {"ok": True, "filename": filename, "web_url": web_url,
            "size_kb": round(len(content) / 1024, 1)}


@router.post("/export/shifts")
def export_shifts_excel(
    db: Session = Depends(get_db),
    _: None = Depends(require_role("supervisor")),
):
    """Export all shift reports to an Excel file in OneDrive."""
    _require_onedrive()
    content = _build_shifts_excel(db)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"shifts_{ts}.xlsx"
    folder = f"{_FOLDER}/exports"
    web_url = od.upload_file(
        folder, filename, content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return {"ok": True, "filename": filename, "web_url": web_url,
            "size_kb": round(len(content) / 1024, 1)}


@router.post("/export/inventory")
def export_inventory_excel(
    db: Session = Depends(get_db),
    _: None = Depends(require_role("warehouse_supervisor")),
):
    """Export stock on hand + transactions to an Excel file in OneDrive."""
    _require_onedrive()
    content = _build_inventory_excel(db)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_{ts}.xlsx"
    folder = f"{_FOLDER}/exports"
    web_url = od.upload_file(
        folder, filename, content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return {"ok": True, "filename": filename, "web_url": web_url,
            "size_kb": round(len(content) / 1024, 1)}


@router.get("/files")
def list_onedrive_files(_: None = Depends(require_role("supervisor"))):
    """List files in the OneDrive production reports folder."""
    _require_onedrive()
    files = od.list_files(_FOLDER)
    return {"folder": _FOLDER, "files": files}


@router.post("/sync/all", response_model=SyncResult)
def sync_all(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    _: None = Depends(require_role("admin")),
):
    """
    Run all OneDrive exports in one call:
      1. DB backup (SQLite only)
      2. Shifts Excel export
      3. Inventory Excel export
    """
    _require_onedrive()
    result = SyncResult(ok=True)
    errors = []

    # 1. DB backup
    try:
        db_url = os.environ.get("DATABASE_URL", "sqlite:///./production_app.db")
        if "sqlite" in db_url:
            db_path = db_url.replace("sqlite:///", "").replace("sqlite://", "")
            candidates = [db_path, "./production_app.db", "production_app.db", "/app/production_app.db"]
            found = next((p for p in candidates if os.path.exists(p)), None)
            if found:
                with open(found, "rb") as f:
                    content = f.read()
                ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                fname = f"production_app_{ts}.db"
                url = od.upload_file(f"{_FOLDER}/backups", fname, content)
                result.db_backup = url
    except Exception as e:
        errors.append(f"DB backup: {e}")

    # 2. Shifts Excel
    try:
        content = _build_shifts_excel(db)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fname = f"shifts_{ts}.xlsx"
        url = od.upload_file(
            f"{_FOLDER}/exports", fname, content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result.shifts_excel = url
    except Exception as e:
        errors.append(f"Shifts export: {e}")

    # 3. Inventory Excel
    try:
        content = _build_inventory_excel(db)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fname = f"inventory_{ts}.xlsx"
        url = od.upload_file(
            f"{_FOLDER}/exports", fname, content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result.inventory_excel = url
    except Exception as e:
        errors.append(f"Inventory export: {e}")

    result.errors = errors
    result.ok = len(errors) == 0
    return result
